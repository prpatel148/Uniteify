import openpyxl
import pandas as pd


# def getCases(self):
#
result = pd.read_excel('C:\\Users\\prati\\OneDrive\\Documents\\GitHub\\Uniteify\\TestData\\RegistrationData.xlsx')
count_row = result.shape[0]
print(result)
print(count_row)
dict = {}
for i in range(count_row):
    r = result.iloc[i]
    print(r)

# book = openpyxl.load_workbook("C:\\Users\\prati\\OneDrive\\Documents\\GitHub\\Uniteify\\TestData\\RegistrationData.xlsx")
# sheet = book.active
# dict = {}
# for i in range(1, sheet.max_row + 1):
#     if sheet.cell(row=i, column=1).value == "test3":
#         for j in range(2, sheet.max_column + 1):
#             dict[(sheet.cell(row=1, column=j).value)] = (sheet.cell(row=i, column=j).value)


